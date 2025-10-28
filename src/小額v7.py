import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl.styles import PatternFill, Alignment


class ExcelAnalyzer:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Excel異常資料分析工具")
        self.window.geometry("500x500")

        # 設定視窗樣式
        self.window.configure(bg='#f0f0f0')

        # 建立標題標籤
        title_label = tk.Label(
            self.window,
            text="Excel 異常資料分析工具",
            font=('Arial', 16, 'bold'),
            bg='#f0f0f0',
            pady=20
        )
        title_label.pack()

        # 建立支出用途關鍵字標籤和輸入框
        purpose_label = tk.Label(
            self.window,
            text="排除支出用途包含:",
            bg='#f0f0f0',
            font=('Arial', 10)
        )
        purpose_label.pack()

        self.purpose_entry = tk.Entry(
            self.window,
            width=40,
            font=('Arial', 10)
        )
        self.purpose_entry.pack(pady=(5, 15))

        # 建立受款人關鍵字標籤和輸入框
        keyword_label = tk.Label(
            self.window,
            text="排除受款人名稱包含:",
            bg='#f0f0f0',
            font=('Arial', 10)
        )
        keyword_label.pack()

        self.keyword_entry = tk.Entry(
            self.window,
            width=40,
            font=('Arial', 10)
        )
        self.keyword_entry.pack(pady=(5, 20))

        # 建立按鈕
        self.select_button = tk.Button(
            self.window,
            text="1. 選擇Excel檔案進行初步分析",
            command=self.analyze_excel,
            width=30,
            height=2,
            bg='#4CAF50',
            fg='white',
            font=('Arial', 10)
        )
        self.select_button.pack(pady=10)

        self.format_button = tk.Button(
            self.window,
            text="2. 選擇分析結果檔案進行進階整理",
            command=self.format_result,
            width=30,
            height=2,
            bg='#2196F3',
            fg='white',
            font=('Arial', 10)
        )
        self.format_button.pack(pady=10)

        self.classify_button = tk.Button(
            self.window,
            text="3. 選擇進階整理結果進行分類分析",
            command=self.classify_result,
            width=30,
            height=2,
            bg='#9C27B0',
            fg='white',
            font=('Arial', 10)
        )
        self.classify_button.pack(pady=10)

        self.status_label = tk.Label(
            self.window,
            text="請選擇要執行的操作",
            bg='#f0f0f0',
            font=('Arial', 9)
        )
        self.status_label.pack(pady=20)

        self.window.mainloop()

    def analyze_excel(self):
        try:
            self.status_label.config(text="正在選擇檔案...")
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )

            if file_path:
                self.status_label.config(text="正在分析...")
                df = pd.read_excel(file_path)

                # 獲取關鍵字
                purpose_keywords = [k.strip() for k in self.purpose_entry.get().split(',') if k.strip()]
                name_keywords = [k.strip() for k in self.keyword_entry.get().split(',') if k.strip()]

                # 排除包含關鍵字的資料
                if purpose_keywords:
                    df = df[~df['支出用途'].str.contains('|'.join(purpose_keywords), na=False)]
                if name_keywords:
                    df = df[~df['受款人名稱'].str.contains('|'.join(name_keywords), na=False)]

                # 計算每個受款人的總金額和筆數
                result = df.groupby(['機關名稱', '營利事業統一編號', '受款人名稱']).agg({
                    '金額': ['count', 'sum']
                }).reset_index()

                # 重新命名欄位
                result.columns = ['機關名稱', '營利事業統一編號', '受款人名稱', '筆數', '加總金額']

                # 篩選金額小於15萬的資料
                result = result[result['加總金額'] < 150000]

                # 排序
                result = result.sort_values(['機關名稱', '營利事業統一編號', '受款人名稱'])

                # 儲存結果
                output_path = os.path.splitext(file_path)[0] + '_分析結果.xlsx'
                try:
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        result.to_excel(writer, index=False)
                    self.status_label.config(text=f"分析完成！結果已儲存至：{os.path.basename(output_path)}")
                    messagebox.showinfo("完成", f"分析完成！\n結果已儲存至：{output_path}")
                except PermissionError:
                    self.status_label.config(text="無法儲存檔案，請確認檔案未被其他程式開啟！")
                    messagebox.showerror("錯誤", "無法儲存檔案，請確認：\n1. 檔案未被其他程式開啟\n2. 您有權限寫入該位置")

        except Exception as e:
            self.status_label.config(text="處理過程發生錯誤！")
            messagebox.showerror("錯誤", f"處理過程發生錯誤：{str(e)}")

    def format_result(self):
        try:
            self.status_label.config(text="正在選擇檔案...")
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )

            if file_path:
                self.status_label.config(text="正在進行進階整理...")
                df = pd.read_excel(file_path)

                # 依照營利事業統一編號分組
                grouped = df.groupby('營利事業統一編號')
                result_data = []

                for tax_id, group in grouped:
                    if len(group) >= 2:  # 只處理有2筆以上的統編
                        result_data.extend(group.to_dict('records'))

                if result_data:
                    result_df = pd.DataFrame(result_data)
                    result_df = result_df.sort_values(['機關名稱', '營利事業統一編號', '受款人名稱'])

                    try:
                        output_path = os.path.splitext(file_path)[0] + '_進階整理.xlsx'
                        if os.path.exists(output_path):
                            try:
                                with open(output_path, 'a'):
                                    pass
                            except PermissionError:
                                base, ext = os.path.splitext(output_path)
                                i = 1
                                while True:
                                    new_path = f"{base}_{i}{ext}"
                                    if not os.path.exists(new_path):
                                        output_path = new_path
                                        break
                                    i += 1

                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            result_df.to_excel(writer, sheet_name='整理結果', index=False)

                        self.status_label.config(text=f"進階整理完成！結果已儲存至：{os.path.basename(output_path)}")
                        messagebox.showinfo("完成", f"進階整理完成！\n結果已儲存至：{output_path}")
                    except PermissionError:
                        self.status_label.config(text="無法儲存檔案，請確認檔案未被其他程式開啟！")
                        messagebox.showerror("錯誤",
                                             "無法儲存檔案，請確認：\n1. 檔案未被其他程式開啟\n2. 您有權限寫入該位置")
                else:
                    self.status_label.config(text="沒有符合條件的資料！")
                    messagebox.showinfo("完成", "沒有符合條件的資料！")

        except Exception as e:
            self.status_label.config(text="處理過程發生錯誤！")
            messagebox.showerror("錯誤", f"處理過程發生錯誤：{str(e)}")

    def classify_result(self):
        try:
            self.status_label.config(text="正在選擇檔案...")
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )

            if file_path:
                self.status_label.config(text="正在進行分類分析...")
                df = pd.read_excel(file_path)

                all_under_150k = []
                mixed_amounts = []

                # 建立兩個計數器來追踪序號
                last_seq_under = 0
                last_seq_mixed = 0
                processed_tax_ids = set()

                # 按照原始資料順序處理
                for _, row in df.iterrows():
                    tax_id = row['營利事業統一編號']
                    if tax_id in processed_tax_ids:
                        continue

                    processed_tax_ids.add(tax_id)
                    tax_df = df[df['營利事業統一編號'] == tax_id]

                    if all(tax_df['加總金額'] < 150000):
                        last_seq_under += 1
                        seq = f"{last_seq_under}."
                        for _, sub_row in tax_df.iterrows():
                            all_under_150k.append({
                                '序號': seq,
                                '機關名稱': sub_row['機關名稱'],
                                '營利事業統一編號': tax_id,
                                '受款人名稱': sub_row['受款人名稱'],
                                '筆數': sub_row['筆數'],
                                '加總金額': sub_row['加總金額']
                            })
                    else:
                        last_seq_mixed += 1
                        seq = f"{last_seq_mixed}."
                        for _, sub_row in tax_df.iterrows():
                            mixed_amounts.append({
                                '序號': seq,
                                '機關名稱': sub_row['機關名稱'],
                                '營利事業統一編號': tax_id,
                                '受款人名稱': sub_row['受款人名稱'],
                                '筆數': sub_row['筆數'],
                                '加總金額': sub_row['加總金額']
                            })

                try:
                    output_path = os.path.splitext(file_path)[0] + '_分類分析.xlsx'
                    if os.path.exists(output_path):
                        try:
                            with open(output_path, 'a'):
                                pass
                        except PermissionError:
                            base, ext = os.path.splitext(output_path)
                            i = 1
                            while True:
                                new_path = f"{base}_{i}{ext}"
                                if not os.path.exists(new_path):
                                    output_path = new_path
                                    break
                                i += 1

                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        def process_sheet(data, sheet_name):
                            if not data:
                                return

                            df_sheet = pd.DataFrame(data)
                            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

                            worksheet = writer.sheets[sheet_name]
                            colors = ['FFFFFFFF', 'FFD9E1F2']  # 白色和淺藍色

                            current_tax_id = None
                            color_index = 0
                            merge_start = 2

                            for row in range(2, len(df_sheet) + 2):
                                tax_id = df_sheet.iloc[row - 2]['營利事業統一編號']

                                if tax_id != current_tax_id or row == len(df_sheet) + 1:
                                    if current_tax_id is not None:
                                        if merge_start < row:
                                            worksheet.merge_cells(f'A{merge_start}:A{row - 1}')
                                            cell = worksheet.cell(row=merge_start, column=1)
                                            cell.alignment = Alignment(horizontal='center', vertical='center')

                                    current_tax_id = tax_id
                                    color_index = (color_index + 1) % 2
                                    merge_start = row

                                for col in range(1, 7):
                                    cell = worksheet.cell(row=row, column=col)
                                    cell.fill = PatternFill(start_color=colors[color_index],
                                                            end_color=colors[color_index],
                                                            fill_type='solid')

                            # 處理最後一組資料的合併
                            if merge_start < len(df_sheet) + 2:
                                worksheet.merge_cells(f'A{merge_start}:A{len(df_sheet) + 1}')
                                cell = worksheet.cell(row=merge_start, column=1)
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                            # 設置欄寬
                            worksheet.column_dimensions['A'].width = 8  # 序號
                            worksheet.column_dimensions['B'].width = 20  # 機關名稱
                            worksheet.column_dimensions['C'].width = 15  # 統編
                            worksheet.column_dimensions['D'].width = 30  # 受款人名稱
                            worksheet.column_dimensions['E'].width = 10  # 筆數
                            worksheet.column_dimensions['F'].width = 15  # 加總金額

                        # 處理兩個sheet
                        process_sheet(all_under_150k, '全部未達15萬')
                        process_sheet(mixed_amounts, '部分達15萬')

                    self.status_label.config(text=f"分類分析完成！結果已儲存至：{os.path.basename(output_path)}")
                    messagebox.showinfo("完成", f"分類分析完成！\n結果已儲存至：{output_path}")
                except PermissionError:
                    self.status_label.config(text="無法儲存檔案，請確認檔案未被其他程式開啟！")
                    messagebox.showerror("錯誤", "無法儲存檔案，請確認：\n1. 檔案未被其他程式開啟\n2. 您有權限寫入該位置")

        except Exception as e:
            self.status_label.config(text="處理過程發生錯誤！")
            messagebox.showerror("錯誤", f"處理過程發生錯誤：{str(e)}")


if __name__ == "__main__":
    app = ExcelAnalyzer()